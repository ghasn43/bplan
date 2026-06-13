"""Excel financial model export — a live, formula-linked annual workbook.

The workbook is NOT a static dump: assumption/driver cells are inputs and every
derived figure is an Excel formula. The model is built so the Balance Sheet
balances by construction:

    NWC          = Receivables + Inventory - Payables
    PPE (NBV)    = cumulative CapEx - cumulative Depreciation
    Retained earn= cumulative Net Profit
    Operating CF = Net Profit + Depreciation - increase in NWC
    Investing CF = -CapEx
    Financing CF = Capital + Loan drawdowns - principal repayments
    Cash (BS)    = Cash Flow closing cash

With those identities, Assets ≡ Equity + Liabilities, so the balance check and
cash reconciliation evaluate to zero in Excel, and the whole model recalculates
when the user edits a driver (units, price, salary, days, rate, ...).
"""
from __future__ import annotations

import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from ..models import BusinessPlanProject
from ..schemas.excel_export import ExcelExportFile
from . import report_data_service as rd

EXPORTS_DIR = Path(__file__).resolve().parent.parent.parent / "generated_exports"

# ---- palette + formats ----------------------------------------------------
NAVY = "1F2A44"
TEAL = "0E7C7B"
SLATE = "475569"
LIGHT = "EAF1FB"
TOTAL_FILL = "E8EEF7"
GREEN = "C6EFCE"
AMBER = "FFEB9C"
RED = "FFC7CE"

CUR = '#,##0;[Red](#,##0);-'
PCT = '0.0%;[Red](0.0%);-'
RATIO = '0.0"x"'
NUM = '#,##0;[Red](#,##0);-'

F_INPUT = Font(color="0000FF")            # blue: user input
F_FORMULA = Font(color="000000")          # black: in-sheet formula
F_LINK = Font(color="008000")             # green: cross-sheet link
F_WARN = Font(color="C00000", bold=True)  # red: warnings
F_LABEL = Font(color="1F2A44")
F_BOLD = Font(bold=True, color="1F2A44")
F_HEAD = Font(bold=True, color="FFFFFF", size=11)
F_TITLE = Font(bold=True, color="FFFFFF", size=20)
F_SUB = Font(color="FFFFFF", size=12)

FILL_HEAD = PatternFill("solid", fgColor=NAVY)
FILL_TEAL = PatternFill("solid", fgColor=TEAL)
FILL_INPUT = PatternFill("solid", fgColor=LIGHT)
FILL_TOTAL = PatternFill("solid", fgColor=TOTAL_FILL)
THIN = Side(style="thin", color="D7DEE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPLINE = Border(top=Side(style="thin", color=NAVY))
UNLOCKED = Protection(locked=False)


def _sanitize(name: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_-]+", "_", name or "").strip("_")
    return name[:60] or "model"


class _Builder:
    def __init__(self, project: BusinessPlanProject, request):
        self.project = project
        self.request = request
        scenario = request.scenario if request.scenario in ("base", "conservative", "optimistic") else "base"
        self.scenario = scenario
        self.ctx = rd.build_report_context(project, scenario, "yearly", _AllOn())
        self.periods = self.ctx["periods"]
        self.n = len(self.periods)
        self.cur = self.ctx["currency"]
        self.wb = Workbook()
        self.refs: dict[str, tuple[str, int]] = {}     # logical -> (sheet, row)
        self.warnings: list[str] = []

    # -- column helpers ---------------------------------------------------
    def col(self, y: int) -> str:
        return get_column_letter(2 + y)

    def ref(self, key: str, y: int) -> str:
        sheet, row = self.refs[key]
        return f"'{sheet}'!{self.col(y)}{row}"

    def last_col(self) -> str:
        return self.col(self.n - 1)

    # -- low-level cell writers ------------------------------------------
    def _hdr(self, ws, row: int, label: str = "Figures in"):
        c = ws.cell(row=row, column=1, value=f"{label} {self.cur} (000s as entered)")
        c.font = Font(italic=True, color=SLATE, size=9)

    def period_header(self, ws, row: int, first_label: str = "Line item"):
        c = ws.cell(row=row, column=1, value=first_label)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
        for y in range(self.n):
            pc = ws.cell(row=row, column=2 + y, value=self.periods[y])
            pc.font = F_HEAD
            pc.fill = FILL_HEAD
            pc.alignment = Alignment(horizontal="right")
            pc.border = BORDER

    def label(self, ws, row, text, *, bold=False, indent=0):
        c = ws.cell(row=row, column=1, value=text)
        c.font = F_BOLD if bold else F_LABEL
        c.alignment = Alignment(indent=indent)
        return c

    def input_row(self, ws, row, label, values, *, fmt=CUR, indent=1):
        self.label(ws, row, label, indent=indent)
        for y in range(self.n):
            c = ws.cell(row=row, column=2 + y, value=float(values[y]) if y < len(values) and values[y] is not None else 0)
            c.font = F_INPUT
            c.fill = FILL_INPUT
            c.number_format = fmt
            c.protection = UNLOCKED
            c.border = BORDER

    def formula_row(self, ws, row, label, fn, *, fmt=CUR, bold=False, link=False, total=False, indent=1):
        self.label(ws, row, label, bold=bold, indent=indent)
        for y in range(self.n):
            c = ws.cell(row=row, column=2 + y, value=fn(y))
            c.font = F_LINK if link else (F_BOLD if bold else F_FORMULA)
            c.number_format = fmt
            c.border = BORDER
            if total:
                c.fill = FILL_TOTAL
                c.border = TOPLINE
        return row

    def section(self, ws, row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF")
        for col in range(1, self.n + 2):
            cc = ws.cell(row=row, column=col)
            cc.fill = FILL_TEAL
        c.font = Font(bold=True, color="FFFFFF")
        return row

    def autosize(self, ws, first=28):
        ws.column_dimensions["A"].width = first
        for y in range(self.n):
            ws.column_dimensions[self.col(y)].width = 15

    def named(self, name, sheet, cell):
        self.wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!{cell}"))


class _AllOn:
    """Options object for build_report_context (always include everything)."""
    scenario = "base"
    view = "yearly"
    report_style = "full"
    include_charts = True
    include_appendices = True
    include_assumptions = True
    include_warnings = True
    include_text_plan = False
    text_plan_include_completed = True
    text_plan_include_drafts = True
    text_plan_include_images = False
    text_plan_include_guidance = False


# ==========================================================================
# public entry point
# ==========================================================================
def generate_excel_financial_model(project: BusinessPlanProject, request) -> ExcelExportFile:
    b = _Builder(project, request)
    wb = b.wb
    wb.remove(wb.active)  # drop default sheet

    # Build in dependency order; schedules first so statements can link to them.
    _control_panel(b)
    _company_project(b)
    _revenue(b)
    _direct_costs(b)
    _opex(b)
    _staffing(b)
    _fixed_assets_and_depreciation(b)
    _working_capital(b)
    _loans(b)
    _income_statement(b)
    _tax_vat(b)
    _cash_flow(b)
    _balance_sheet(b)
    _ratios(b)
    if request.include_scenarios:
        _scenario_comparison(b)
    if request.include_charts:
        _charts(b)
    if request.include_checks:
        _checks(b)
    _assumptions_summary(b)
    _data_dictionary(b)
    _cover(b)  # last so it can move to front

    # Cover first, Control Panel second.
    wb.move_sheet("Cover", -wb.sheetnames.index("Cover"))

    _named_ranges(b)
    if request.protect_formulas:
        _protect(b)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    export_id = uuid.uuid4().hex
    base = _sanitize(f"financial_model_{b.ctx['meta']['company']}_{b.ctx['meta']['project_name']}_{b.scenario}_{datetime.now():%Y%m%d}")
    file_name = f"{base}_{export_id[:6]}.xlsx"
    proj_dir = EXPORTS_DIR / _sanitize(project.id)
    proj_dir.mkdir(parents=True, exist_ok=True)
    path = proj_dir / file_name
    wb.save(path)

    entry = ExcelExportFile(
        export_id=export_id, project_id=project.id, file_name=file_name,
        file_size=path.stat().st_size, scenario=b.scenario, status="ready",
        generated_at=datetime.now(timezone.utc), warnings=b.warnings,
    )
    _add_index_entry(project.id, entry)
    return entry


# ==========================================================================
# index (download registry)
# ==========================================================================
import json  # noqa: E402


def _index_path(project_id: str) -> Path:
    d = EXPORTS_DIR / _sanitize(project_id)
    d.mkdir(parents=True, exist_ok=True)
    return d / "_index.json"


def load_index(project_id: str) -> list[dict]:
    p = _index_path(project_id)
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []


def _add_index_entry(project_id: str, entry: ExcelExportFile) -> None:
    entries = load_index(project_id)
    entries.insert(0, entry.model_dump(mode="json"))
    _index_path(project_id).write_text(json.dumps(entries, indent=2, default=str), encoding="utf-8")


def export_path(project_id: str, file_name: str) -> Path:
    return EXPORTS_DIR / _sanitize(project_id) / file_name


SHEET_LIST = [
    "Cover", "Control Panel", "Company & Project", "Revenue Projection",
    "Direct Cost Projection", "Operating Expense Projection", "Staffing",
    "Depreciation Schedule", "Working Capital", "Loan Schedule",
    "Income Statement", "Tax & VAT", "Cash Flow", "Balance Sheet",
    "Financial Ratios", "Scenario Comparison", "Charts", "Checks",
    "Assumptions Summary", "Data Dictionary",
]


def build_preview(project: BusinessPlanProject, request):
    from ..schemas.excel_export import ExcelExportPreview
    from .completion import build_completion_report
    setup = project.setup
    scenario = request.scenario if request.scenario in ("base", "conservative", "optimistic") else "base"
    labels = {"base": "Base Case", "conservative": "Conservative Case", "optimistic": "Optimistic Case"}
    start = setup.projection_start_date.year if setup and setup.projection_start_date else datetime.now().year
    years = setup.projection_period.years if setup and setup.projection_period else 5
    blocking = []
    if setup is None:
        blocking.append("Project setup is required before exporting the model.")
    if not project.products:
        blocking.append("At least one product / revenue stream is required.")
    warnings = []
    if not project.fixed_assets:
        warnings.append("No fixed assets — the depreciation schedule will be empty.")
    if not (project.financing and project.financing.loans):
        warnings.append("No loans — the loan schedule will be empty.")
    completion = build_completion_report(project)
    if completion.completion_percent < 100:
        warnings.append(f"Project is {completion.completion_percent}% complete; the model reflects current assumptions.")
    return ExcelExportPreview(
        project_id=project.id,
        company_name=(setup.business_name if setup else project.name),
        project_name=(setup.project_name if setup else ""),
        currency=setup.currency if setup else "USD",
        period_range=f"{start}–{start + years - 1}",
        scenario=scenario, scenario_label=labels[scenario],
        sheets=SHEET_LIST, protect_formulas=request.protect_formulas,
        data_ready=len(blocking) == 0, can_generate=len(blocking) == 0,
        blocking=blocking, warnings=warnings, estimated_size_kb=45,
    )


def delete_export(project_id: str, export_id: str) -> bool:
    entries = load_index(project_id)
    entry = next((e for e in entries if e.get("export_id") == export_id), None)
    if not entry:
        return False
    f = export_path(project_id, entry["file_name"])
    if f.exists():
        f.unlink()
    _index_path(project_id).write_text(
        json.dumps([e for e in entries if e.get("export_id") != export_id], indent=2, default=str), encoding="utf-8")
    return True


# ==========================================================================
# sheets
# ==========================================================================
def _cover(b: _Builder):
    ws = b.wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    m = b.ctx["meta"]
    for r in range(1, 16):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFFFFF")
    ws.merge_cells("B2:H4")
    band = ws.cell(row=2, column=2)
    band.fill = FILL_HEAD
    for r in (2, 3, 4):
        for c in range(2, 9):
            ws.cell(row=r, column=c).fill = FILL_HEAD
    t = ws.cell(row=2, column=2, value=m["company"])
    t.font = F_TITLE
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.cell(row=5, column=2, value=m["project_name"]).font = Font(size=13, color=TEAL, bold=True)
    ws.cell(row=7, column=2, value="Integrated Financial Projection Model").font = Font(size=15, bold=True, color=NAVY)
    rows = [
        ("Projection Period", m["period_range"]),
        ("Currency", b.cur),
        ("Scenario", m["scenario_label"]),
        ("Prepared for", m["prepared_for"]),
        ("Generated", m["prepared_date"]),
        ("Application", "Business Plan Studio"),
        ("Workbook version", "1.0"),
    ]
    for i, (k, v) in enumerate(rows):
        ws.cell(row=9 + i, column=2, value=k).font = Font(bold=True, color=SLATE)
        ws.cell(row=9 + i, column=3, value=v).font = F_LABEL
    conf = ws.cell(row=18, column=2, value="CONFIDENTIAL")
    conf.font = Font(bold=True, color="C00000")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 40


def _control_panel(b: _Builder):
    ws = b.wb.create_sheet("Control Panel")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Control Panel").font = Font(bold=True, size=14, color=NAVY)
    ws.cell(row=2, column=1, value="Blue = input · Black = formula · Green = link · Red = warning").font = Font(italic=True, color=SLATE, size=9)
    setup = b.project.setup
    tax = b.project.tax
    wc = b.project.working_capital
    rows = [
        ("Selected Scenario", b.ctx["meta"]["scenario_label"], "list"),
        ("Currency", b.cur, None),
        ("Projection Start", (setup.projection_start_date.isoformat() if setup and setup.projection_start_date else ""), None),
        ("Projection Years", b.n, None),
        ("Corporate Tax Rate", _frac(getattr(tax, "corporate_tax_rate", 0)), PCT),
        ("VAT Rate", _frac(getattr(tax, "vat_rate", 0)), PCT),
        ("Minimum Cash Balance", float(getattr(wc, "minimum_cash_balance", 0) or 0), CUR),
    ]
    r0 = 4
    keymap = {}
    for i, (k, v, kind) in enumerate(rows):
        r = r0 + i
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, color=SLATE)
        c = ws.cell(row=r, column=2, value=v)
        c.font = F_INPUT
        c.fill = FILL_INPUT
        c.protection = UNLOCKED
        c.border = BORDER
        if kind == PCT:
            c.number_format = PCT
        elif kind == CUR:
            c.number_format = CUR
        keymap[k] = r
    # scenario dropdown
    dv = DataValidation(type="list", formula1='"Base Case,Conservative Case,Optimistic Case"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=keymap["Selected Scenario"], column=2))
    b.refs["ctrl_scenario"] = ("Control Panel", keymap["Selected Scenario"])
    b.refs["ctrl_tax"] = ("Control Panel", keymap["Corporate Tax Rate"])
    b.refs["ctrl_vat"] = ("Control Panel", keymap["VAT Rate"])
    b.refs["ctrl_mincash"] = ("Control Panel", keymap["Minimum Cash Balance"])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22


def _frac(v) -> float:
    try:
        v = float(v or 0)
    except (TypeError, ValueError):
        return 0.0
    return v / 100.0 if v > 1 else v


def _company_project(b: _Builder):
    ws = b.wb.create_sheet("Company & Project")
    ws.sheet_view.showGridLines = False
    setup = b.project.setup
    ws.cell(row=1, column=1, value="Company").font = Font(bold=True, size=13, color=NAVY)
    company_rows = [
        ("Company name", b.ctx["meta"]["company"]),
        ("Industry", getattr(setup, "industry", "") if setup else ""),
        ("Country", getattr(setup, "country", "") if setup else ""),
        ("City", getattr(setup, "city", "") if setup else ""),
        ("Description", getattr(setup, "business_description", "") if setup else ""),
    ]
    r = 2
    for k, v in company_rows:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, color=SLATE)
        ws.cell(row=r, column=2, value=v).font = F_LABEL
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Project").font = Font(bold=True, size=13, color=NAVY)
    r += 1
    project_rows = [
        ("Project name", b.ctx["meta"]["project_name"]),
        ("Projection start", (setup.projection_start_date.isoformat() if setup and setup.projection_start_date else "")),
        ("Projection period", b.ctx["meta"]["period_range"]),
        ("Reporting standard", getattr(getattr(setup, "reporting_standard", None), "value", "") if setup else ""),
        ("Currency", b.cur),
    ]
    for k, v in project_rows:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, color=SLATE)
        ws.cell(row=r, column=2, value=v).font = F_LABEL
        r += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def _cat_sheet(b, title, table, total_key, total_label):
    """Generic schedule: input rows per category + a SUM total row."""
    ws = b.wb.create_sheet(title)
    b._hdr(ws, 1)
    hdr = 3
    b.period_header(ws, hdr)
    r = hdr + 1
    first = r
    for row in table:
        b.input_row(ws, r, row["label"], row.get("values", []))
        r += 1
    last = r - 1
    total_row = r
    b.refs[total_key] = (title, total_row)

    def fn(y):
        cl = b.col(y)
        return f"=SUM({cl}{first}:{cl}{last})" if last >= first else 0
    b.formula_row(ws, total_row, total_label, fn, bold=True, total=True, indent=0)
    ws.freeze_panes = f"B{hdr + 1}"
    ws.auto_filter.ref = f"A{hdr}:{b.last_col()}{last if last >= first else hdr}"
    b.autosize(ws, first=34)
    return ws


def _revenue(b: _Builder):
    ws = _cat_sheet(b, "Revenue Projection", b.ctx["revenue_table"], "rev_total", "Total Net Revenue")
    tr = b.refs["rev_total"][1]
    b.formula_row(ws, tr + 1, "VAT Output (memo)", lambda y: f"={b.col(y)}{tr}*VAT_Rate", link=True, indent=0)


def _direct_costs(b: _Builder):
    _cat_sheet(b, "Direct Cost Projection", b.ctx["direct_cost_table"], "dc_total", "Total Cost of Sales")


def _opex(b: _Builder):
    _cat_sheet(b, "Operating Expense Projection", b.ctx["opex_table"], "opex_total", "Total Operating Expenses")


def _staffing(b: _Builder):
    table = b.ctx["staff_by_dept"] or [{"label": "Total staff cost", "values": _staff_fallback(b)}]
    _cat_sheet(b, "Staffing", table, "staff_total", "Total Staff Cost")


def _staff_fallback(b):
    # If no per-department breakdown, derive from operating expenses staff line.
    return [0.0] * b.n


def _fixed_assets_and_depreciation(b: _Builder):
    ws = b.wb.create_sheet("Depreciation Schedule")
    b._hdr(ws, 1)
    assets = b.project.fixed_assets
    setup = b.project.setup
    start_year = setup.projection_start_date.year if setup and setup.projection_start_date else datetime.now().year

    # Asset register (inputs) at top
    reg_hdr = 3
    for j, h in enumerate(["Asset", "Purchase cost", "Residual", "Life (yrs)", "Purchase year"]):
        c = ws.cell(row=reg_hdr, column=1 + j, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
    asset_rows = {}
    r = reg_hdr + 1
    for a in assets:
        py = (a.purchase_date.year - start_year) if getattr(a, "purchase_date", None) else 0
        py = max(0, min(py, b.n - 1))
        ws.cell(row=r, column=1, value=a.name).font = F_LABEL
        for col, val in ((2, float(a.purchase_amount or 0)), (3, float(a.residual_value or 0)),
                         (4, float(a.useful_life_years or 1)), (5, py + 1)):
            cc = ws.cell(row=r, column=col, value=val)
            cc.font = F_INPUT
            cc.fill = FILL_INPUT
            cc.protection = UNLOCKED
            cc.number_format = CUR if col in (2, 3) else NUM
            cc.border = BORDER
        asset_rows[r] = py
        r += 1
    reg_last = r - 1

    # Annual additions + depreciation + NBV (formulas)
    grid_hdr = r + 1
    b.period_header(ws, grid_hdr)
    add_row = grid_hdr + 1
    dep_row = grid_hdr + 2
    nbv_row = grid_hdr + 3

    # Additions: sum of purchase costs whose purchase-year column == this year
    def add_fn(y):
        terms = []
        for rr, py in asset_rows.items():
            if py == y:
                terms.append(f"B{rr}")
        return "=" + ("+".join(terms) if terms else "0")
    b.formula_row(ws, add_row, "CapEx additions", add_fn, bold=True, total=True, indent=0)

    # Depreciation: straight-line per asset, only after purchase year
    def dep_fn(y):
        terms = []
        for rr, py in asset_rows.items():
            if y >= py:
                terms.append(f"MAX(0,(B{rr}-C{rr})/D{rr})")
        return "=" + ("+".join(terms) if terms else "0")
    b.formula_row(ws, dep_row, "Depreciation charge", dep_fn, bold=True, total=True, indent=0)

    # Closing NBV = cumulative additions - cumulative depreciation
    def nbv_fn(y):
        ac = b.col(y)
        if y == 0:
            return f"={ac}{add_row}-{ac}{dep_row}"
        pc = b.col(y - 1)
        return f"={pc}{nbv_row}+{ac}{add_row}-{ac}{dep_row}"
    b.formula_row(ws, nbv_row, "Closing net book value", nbv_fn, bold=True, total=True, indent=0)

    b.refs["capex"] = ("Depreciation Schedule", add_row)
    b.refs["dep"] = ("Depreciation Schedule", dep_row)
    b.refs["nbv"] = ("Depreciation Schedule", nbv_row)
    b.autosize(ws, first=30)
    ws.column_dimensions["E"].width = 14


def _working_capital(b: _Builder):
    ws = b.wb.create_sheet("Working Capital")
    b._hdr(ws, 1)
    hdr = 3
    b.period_header(ws, hdr)
    setup_wc = b.project.working_capital
    ar_days = float(getattr(setup_wc, "accounts_receivable_days", 0) or 0)
    inv_days = float(getattr(setup_wc, "inventory_days", 0) or 0)
    ap_days = float(getattr(setup_wc, "accounts_payable_days", 0) or 0)
    credit = _frac(getattr(setup_wc, "percent_sales_on_credit", 100) or 100)

    r = hdr + 1
    # day assumptions (inputs)
    b.input_row(ws, r, "Receivable days", [ar_days] * b.n, fmt=NUM)
    ar_d = r
    r += 1
    b.input_row(ws, r, "Inventory days", [inv_days] * b.n, fmt=NUM)
    inv_d = r
    r += 1
    b.input_row(ws, r, "Payable days", [ap_days] * b.n, fmt=NUM)
    ap_d = r
    r += 1
    b.input_row(ws, r, "% sales on credit", [credit] * b.n, fmt=PCT)
    cr_d = r
    r += 1
    # balances (formulas)
    rec_row = r
    b.formula_row(ws, rec_row, "Trade receivables", lambda y: f"={b.ref('rev_total', y)}*{b.col(y)}{cr_d}*{b.col(y)}{ar_d}/365", link=True)
    r += 1
    inv_row = r
    b.formula_row(ws, inv_row, "Inventory", lambda y: f"={b.ref('dc_total', y)}*{b.col(y)}{inv_d}/365", link=True)
    r += 1
    pay_row = r
    b.formula_row(ws, pay_row, "Trade payables", lambda y: f"={b.ref('dc_total', y)}*{b.col(y)}{ap_d}/365", link=True)
    r += 1
    nwc_row = r
    b.formula_row(ws, nwc_row, "Net working capital", lambda y: f"={b.col(y)}{rec_row}+{b.col(y)}{inv_row}-{b.col(y)}{pay_row}", bold=True, total=True, indent=0)
    r += 1
    mv_row = r
    b.formula_row(ws, mv_row, "Increase in NWC", lambda y: (f"={b.col(y)}{nwc_row}" if y == 0 else f"={b.col(y)}{nwc_row}-{b.col(y-1)}{nwc_row}"), bold=True, indent=0)
    b.refs["wc_rec"] = ("Working Capital", rec_row)
    b.refs["wc_inv"] = ("Working Capital", inv_row)
    b.refs["wc_pay"] = ("Working Capital", pay_row)
    b.refs["wc_nwc"] = ("Working Capital", nwc_row)
    b.refs["wc_move"] = ("Working Capital", mv_row)
    ws.freeze_panes = f"B{hdr + 1}"
    b.autosize(ws, first=30)


def _loans(b: _Builder):
    ws = b.wb.create_sheet("Loan Schedule")
    b._hdr(ws, 1)
    loans = b.project.financing.loans if b.project.financing else []
    hdr = 3
    b.period_header(ws, hdr)
    r = hdr + 1
    interest_rows, closing_rows, draw_rows, repay_rows = [], [], [], []
    for i, loan in enumerate(loans):
        amount = float(loan.amount or 0)
        rate = _frac(getattr(loan, "interest_rate", 0))
        term_y = max(1, int(round((getattr(loan, "repayment_period_months", 12) or 12) / 12)))
        b.section(ws, r, f"Loan {i + 1}: {loan.name}")
        r += 1
        # drawdown (year 0), repayment = principal/term after grace (simplified equal principal)
        draw = [amount if y == 0 else 0 for y in range(b.n)]
        b.input_row(ws, r, "Drawdown", draw)
        draw_row = r
        draw_rows.append(draw_row)
        r += 1
        b.input_row(ws, r, "Annual interest rate", [rate] * b.n, fmt=PCT)
        rate_row = r
        r += 1
        open_row = r
        # Opening = prior year's closing (closing is 3 rows below: repay, interest, closing).
        b.formula_row(ws, open_row, "Opening principal",
                      lambda y, open_row=open_row: ("=0" if y == 0 else f"={b.col(y-1)}{open_row + 3}"))
        r += 1
        repay_row = r
        per = round(amount / term_y, 2)

        def repay_fn(y, open_row=open_row, draw_row=draw_row, per=per, term_y=term_y):
            # repay equal principal in years 1..term (after the drawdown year 0)
            if 1 <= y <= term_y:
                return f"=MIN({per},{b.col(y)}{open_row}+{b.col(y)}{draw_row})"
            return "=0"
        b.formula_row(ws, repay_row, "Principal repayment", repay_fn)
        repay_rows.append(repay_row)
        r += 1
        int_row = r
        b.formula_row(ws, int_row, "Interest expense",
                      lambda y, open_row=open_row, draw_row=draw_row, rate_row=rate_row: f"=({b.col(y)}{open_row}+{b.col(y)}{draw_row})*{b.col(y)}{rate_row}")
        interest_rows.append(int_row)
        r += 1
        close_row = r
        b.formula_row(ws, close_row, "Closing principal",
                      lambda y, open_row=open_row, draw_row=draw_row, repay_row=repay_row: f"={b.col(y)}{open_row}+{b.col(y)}{draw_row}-{b.col(y)}{repay_row}", bold=True, total=True, indent=0)
        closing_rows.append(close_row)
        r += 2

    # Totals
    b.section(ws, r, "Totals")
    r += 1
    tot_int = r
    b.formula_row(ws, tot_int, "Total interest expense", lambda y: "=" + ("+".join(f"{b.col(y)}{ir}" for ir in interest_rows) if interest_rows else "0"), bold=True, total=True, indent=0)
    r += 1
    tot_close = r
    b.formula_row(ws, tot_close, "Total closing borrowings", lambda y: "=" + ("+".join(f"{b.col(y)}{cr}" for cr in closing_rows) if closing_rows else "0"), bold=True, total=True, indent=0)
    r += 1
    tot_draw = r
    b.formula_row(ws, tot_draw, "Total drawdowns", lambda y: "=" + ("+".join(f"{b.col(y)}{dr}" for dr in draw_rows) if draw_rows else "0"), indent=0)
    r += 1
    tot_repay = r
    b.formula_row(ws, tot_repay, "Total principal repayments", lambda y: "=" + ("+".join(f"{b.col(y)}{rr}" for rr in repay_rows) if repay_rows else "0"), indent=0)
    b.refs["loan_int"] = ("Loan Schedule", tot_int)
    b.refs["loan_close"] = ("Loan Schedule", tot_close)
    b.refs["loan_draw"] = ("Loan Schedule", tot_draw)
    b.refs["loan_repay"] = ("Loan Schedule", tot_repay)
    ws.freeze_panes = f"B{hdr + 1}"
    b.autosize(ws, first=30)


def _tax_vat(b: _Builder):
    ws = b.wb.create_sheet("Tax & VAT")
    b._hdr(ws, 1)
    hdr = 3
    b.period_header(ws, hdr)
    r = hdr + 1
    pbt_row = r
    b.formula_row(ws, pbt_row, "Profit before tax", lambda y: f"={b.ref('is_pbt', y)}", link=True)
    r += 1
    rate_row = r
    b.formula_row(ws, rate_row, "Corporate tax rate", lambda y: "=Corporate_Tax_Rate", fmt=PCT, link=True)
    r += 1
    tax_row = r
    b.formula_row(ws, tax_row, "Tax expense", lambda y: f"=MAX(0,{b.col(y)}{pbt_row})*{b.col(y)}{rate_row}", bold=True, total=True, indent=0)
    r += 2
    vo_row = r
    b.formula_row(ws, vo_row, "VAT output", lambda y: f"={b.ref('rev_total', y)}*VAT_Rate", link=True)
    r += 1
    vi_row = r
    b.formula_row(ws, vi_row, "VAT input", lambda y: f"=({b.ref('dc_total', y)}+{b.ref('opex_total', y)})*VAT_Rate", link=True)
    r += 1
    vn_row = r
    b.formula_row(ws, vn_row, "Net VAT position", lambda y: f"={b.col(y)}{vo_row}-{b.col(y)}{vi_row}", bold=True, total=True, indent=0)
    b.refs["tax_exp"] = ("Tax & VAT", tax_row)
    b.autosize(ws, first=30)


def _income_statement(b: _Builder):
    ws = b.wb.create_sheet("Income Statement")
    b._hdr(ws, 1)
    hdr = 3
    b.period_header(ws, hdr)
    r = hdr + 1
    rev = r
    b.formula_row(ws, rev, "Revenue", lambda y: f"={b.ref('rev_total', y)}", link=True, bold=True)
    r += 1
    cogs = r
    b.formula_row(ws, cogs, "Cost of sales", lambda y: f"=-{b.ref('dc_total', y)}", link=True)
    r += 1
    gp = r
    b.formula_row(ws, gp, "Gross profit", lambda y: f"={b.col(y)}{rev}+{b.col(y)}{cogs}", bold=True, total=True, indent=0)
    r += 1
    opex = r
    b.formula_row(ws, opex, "Operating expenses", lambda y: f"=-{b.ref('opex_total', y)}", link=True)
    r += 1
    staff = r
    b.formula_row(ws, staff, "Staff costs", lambda y: f"=-{b.ref('staff_total', y)}", link=True)
    r += 1
    dep = r
    b.formula_row(ws, dep, "Depreciation", lambda y: f"=-{b.ref('dep', y)}", link=True)
    r += 1
    op = r
    b.formula_row(ws, op, "Operating profit", lambda y: f"={b.col(y)}{gp}+{b.col(y)}{opex}+{b.col(y)}{staff}+{b.col(y)}{dep}", bold=True, total=True, indent=0)
    r += 1
    fin = r
    b.formula_row(ws, fin, "Finance costs", lambda y: f"=-{b.ref('loan_int', y)}", link=True)
    r += 1
    pbt = r
    b.formula_row(ws, pbt, "Profit before tax", lambda y: f"={b.col(y)}{op}+{b.col(y)}{fin}", bold=True, total=True, indent=0)
    b.refs["is_pbt"] = ("Income Statement", pbt)
    r += 1
    tax = r
    # Self-contained so there is no circular reference with the Tax sheet.
    b.formula_row(ws, tax, "Income tax expense", lambda y: f"=-MAX(0,{b.col(y)}{pbt})*Corporate_Tax_Rate")
    b.refs["is_tax"] = ("Income Statement", tax)
    r += 1
    npr = r
    b.formula_row(ws, npr, "Profit for the period", lambda y: f"={b.col(y)}{pbt}+{b.col(y)}{tax}", bold=True, total=True, indent=0)
    r += 2
    # margins + EBITDA
    b.formula_row(ws, r, "Gross margin %", lambda y: f"=IF({b.col(y)}{rev}=0,0,{b.col(y)}{gp}/{b.col(y)}{rev})", fmt=PCT, indent=0)
    r += 1
    ebitda = r
    b.formula_row(ws, ebitda, "EBITDA", lambda y: f"={b.col(y)}{op}-{b.col(y)}{dep}", bold=True, indent=0)
    r += 1
    b.formula_row(ws, r, "Net profit margin %", lambda y: f"=IF({b.col(y)}{rev}=0,0,{b.col(y)}{npr}/{b.col(y)}{rev})", fmt=PCT, indent=0)
    b.refs["is_rev"] = ("Income Statement", rev)
    b.refs["is_gp"] = ("Income Statement", gp)
    b.refs["is_op"] = ("Income Statement", op)
    b.refs["is_ebitda"] = ("Income Statement", ebitda)
    b.refs["is_np"] = ("Income Statement", npr)
    ws.freeze_panes = f"B{hdr + 1}"
    b.autosize(ws, first=30)


def _cash_flow(b: _Builder):
    ws = b.wb.create_sheet("Cash Flow")
    b._hdr(ws, 1)
    hdr = 3
    b.period_header(ws, hdr)
    eq = b.project.financing.equity if b.project.financing else None
    capital0 = float((getattr(eq, "founder_capital", 0) or 0) + (getattr(eq, "investor_equity", 0) or 0)) if eq else 0.0
    r = hdr + 1
    b.section(ws, r, "Operating activities")
    r += 1
    npr = r
    b.formula_row(ws, npr, "Profit for the period", lambda y: f"={b.ref('is_np', y)}", link=True)
    r += 1
    depr = r
    b.formula_row(ws, depr, "Add: depreciation", lambda y: f"={b.ref('dep', y)}", link=True)
    r += 1
    wcm = r
    b.formula_row(ws, wcm, "Less: increase in NWC", lambda y: f"=-{b.ref('wc_move', y)}", link=True)
    r += 1
    opcf = r
    b.formula_row(ws, opcf, "Net cash from operating", lambda y: f"={b.col(y)}{npr}+{b.col(y)}{depr}+{b.col(y)}{wcm}", bold=True, total=True, indent=0)
    r += 1
    b.section(ws, r, "Investing activities")
    r += 1
    capex = r
    b.formula_row(ws, capex, "Purchase of PPE", lambda y: f"=-{b.ref('capex', y)}", link=True)
    r += 1
    invcf = r
    b.formula_row(ws, invcf, "Net cash from investing", lambda y: f"={b.col(y)}{capex}", bold=True, total=True, indent=0)
    r += 1
    b.section(ws, r, "Financing activities")
    r += 1
    cap = r
    b.input_row(ws, cap, "Capital injections", [capital0 if y == 0 else 0 for y in range(b.n)])
    b.refs["cf_capital"] = ("Cash Flow", cap)
    r += 1
    drw = r
    b.formula_row(ws, drw, "Loan drawdowns", lambda y: f"={b.ref('loan_draw', y)}", link=True)
    r += 1
    rep = r
    b.formula_row(ws, rep, "Principal repayments", lambda y: f"=-{b.ref('loan_repay', y)}", link=True)
    r += 1
    fincf = r
    b.formula_row(ws, fincf, "Net cash from financing", lambda y: f"={b.col(y)}{cap}+{b.col(y)}{drw}+{b.col(y)}{rep}", bold=True, total=True, indent=0)
    r += 1
    opening = r
    b.formula_row(ws, opening, "Opening cash", lambda y: ("=0" if y == 0 else f"={b.col(y-1)}{opening + 1}"))
    r += 1
    closing = r
    b.formula_row(ws, closing, "Closing cash", lambda y: f"={b.col(y)}{opening}+{b.col(y)}{opcf}+{b.col(y)}{invcf}+{b.col(y)}{fincf}", bold=True, total=True, indent=0)
    b.refs["cf_close"] = ("Cash Flow", closing)
    b.refs["cf_op"] = ("Cash Flow", opcf)
    b.refs["cf_inv"] = ("Cash Flow", invcf)
    b.refs["cf_fin"] = ("Cash Flow", fincf)
    ws.freeze_panes = f"B{hdr + 1}"
    b.autosize(ws, first=30)


def _balance_sheet(b: _Builder):
    ws = b.wb.create_sheet("Balance Sheet")
    b._hdr(ws, 1)
    hdr = 3
    b.period_header(ws, hdr)
    r = hdr + 1
    b.section(ws, r, "Assets")
    r += 1
    ppe = r
    b.formula_row(ws, ppe, "Property, plant & equipment", lambda y: f"={b.ref('nbv', y)}", link=True)
    r += 1
    inv = r
    b.formula_row(ws, inv, "Inventory", lambda y: f"={b.ref('wc_inv', y)}", link=True)
    r += 1
    rec = r
    b.formula_row(ws, rec, "Trade receivables", lambda y: f"={b.ref('wc_rec', y)}", link=True)
    r += 1
    cash = r
    b.formula_row(ws, cash, "Cash and cash equivalents", lambda y: f"={b.ref('cf_close', y)}", link=True)
    r += 1
    ta = r
    b.formula_row(ws, ta, "Total assets", lambda y: f"=SUM({b.col(y)}{ppe}:{b.col(y)}{cash})", bold=True, total=True, indent=0)
    b.refs["bs_assets"] = ("Balance Sheet", ta)
    r += 2
    b.section(ws, r, "Equity & liabilities")
    r += 1
    sc = r
    b.formula_row(ws, sc, "Share capital", lambda y: (f"={b.ref('cf_capital', 0)}" if y == 0 else f"={b.col(y-1)}{sc}+{b.ref('cf_capital', y)}"), link=True)
    r += 1
    re = r
    b.formula_row(ws, re, "Retained earnings", lambda y: (f"={b.ref('is_np', 0)}" if y == 0 else f"={b.col(y-1)}{re}+{b.ref('is_np', y)}"), link=True)
    r += 1
    teq = r
    b.formula_row(ws, teq, "Total equity", lambda y: f"={b.col(y)}{sc}+{b.col(y)}{re}", bold=True, indent=0)
    r += 1
    bor = r
    b.formula_row(ws, bor, "Borrowings", lambda y: f"={b.ref('loan_close', y)}", link=True)
    r += 1
    pay = r
    b.formula_row(ws, pay, "Trade payables", lambda y: f"={b.ref('wc_pay', y)}", link=True)
    r += 1
    tel = r
    b.formula_row(ws, tel, "Total equity & liabilities", lambda y: f"={b.col(y)}{teq}+{b.col(y)}{bor}+{b.col(y)}{pay}", bold=True, total=True, indent=0)
    b.refs["bs_eql"] = ("Balance Sheet", tel)
    r += 1
    chk = r
    b.formula_row(ws, chk, "Balance check (=0)", lambda y: f"={b.col(y)}{ta}-{b.col(y)}{tel}", bold=True, indent=0)
    b.refs["bs_cash"] = ("Balance Sheet", cash)
    b.refs["bs_check"] = ("Balance Sheet", chk)
    # conditional format the check row green when zero, red otherwise
    rng = f"{b.col(0)}{chk}:{b.last_col()}{chk}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="notEqual", formula=["0"], fill=PatternFill("solid", fgColor=RED)))
    ws.freeze_panes = f"B{hdr + 1}"
    b.autosize(ws, first=30)


def _ratios(b: _Builder):
    ws = b.wb.create_sheet("Financial Ratios")
    b._hdr(ws, 1)
    hdr = 3
    b.period_header(ws, hdr)
    r = hdr + 1
    rows = [
        ("Gross margin %", lambda y: f"=IF({b.ref('is_rev', y)}=0,0,{b.ref('is_gp', y)}/{b.ref('is_rev', y)})", PCT),
        ("EBITDA margin %", lambda y: f"=IF({b.ref('is_rev', y)}=0,0,{b.ref('is_ebitda', y)}/{b.ref('is_rev', y)})", PCT),
        ("Net profit margin %", lambda y: f"=IF({b.ref('is_rev', y)}=0,0,{b.ref('is_np', y)}/{b.ref('is_rev', y)})", PCT),
        ("Return on assets %", lambda y: f"=IF({b.ref('bs_assets', y)}=0,0,{b.ref('is_np', y)}/{b.ref('bs_assets', y)})", PCT),
        ("Debt to assets", lambda y: f"=IF({b.ref('bs_assets', y)}=0,0,{b.ref('loan_close', y)}/{b.ref('bs_assets', y)})", RATIO),
        ("Interest coverage", lambda y: f"=IF({b.ref('loan_int', y)}=0,0,{b.ref('is_op', y)}/{b.ref('loan_int', y)})", RATIO),
        ("Net working capital", lambda y: f"={b.ref('wc_nwc', y)}", CUR),
        ("Closing cash", lambda y: f"={b.ref('cf_close', y)}", CUR),
    ]
    for label, fn, fmt in rows:
        b.formula_row(ws, r, label, fn, fmt=fmt, link=True, indent=0)
        r += 1
    ws.freeze_panes = f"B{hdr + 1}"
    b.autosize(ws, first=30)


def _scenario_comparison(b: _Builder):
    ws = b.wb.create_sheet("Scenario Comparison")
    ws.cell(row=1, column=1, value="Scenario Comparison (snapshot from the app)").font = Font(bold=True, color=NAVY, size=12)
    sc = b.ctx["scenario_comparison"]
    headers = ["Metric", "Base Case", "Conservative Case", "Optimistic Case", "Variance to Base"]
    for j, h in enumerate(headers):
        c = ws.cell(row=3, column=1 + j, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
    r = 4
    for row in sc["rows"]:
        ws.cell(row=r, column=1, value=row["label"]).font = F_LABEL
        vals = row["values"]
        base = vals.get("base")
        for j, key in enumerate(("base", "conservative", "optimistic")):
            c = ws.cell(row=r, column=2 + j, value=vals.get(key))
            c.number_format = PCT if row["format"] == "percent" else CUR
            c.border = BORDER
        var = ws.cell(row=r, column=5, value=f"=C{r}-B{r}")
        var.number_format = PCT if row["format"] == "percent" else CUR
        r += 1
    for col, w in (("A", 30), ("B", 16), ("C", 18), ("D", 18), ("E", 16)):
        ws.column_dimensions[col].width = w


def _charts(b: _Builder):
    ws = b.wb.create_sheet("Charts")
    ws.sheet_view.showGridLines = False
    isheet = "Income Statement"
    is_rev = b.refs["is_rev"][1]
    is_gp = b.refs["is_gp"][1]
    is_np = b.refs["is_np"][1]
    cats = Reference(b.wb[isheet], min_col=2, max_col=1 + b.n, min_row=3, max_row=3)

    line = LineChart()
    line.title = "Revenue, Gross Profit & Net Profit"
    for row in (is_rev, is_gp, is_np):
        data = Reference(b.wb[isheet], min_col=1, max_col=1 + b.n, min_row=row, max_row=row)
        line.add_data(data, titles_from_data=True, from_rows=True)
    line.set_categories(cats)
    line.height = 8
    line.width = 18
    ws.add_chart(line, "B2")

    bar = BarChart()
    bar.title = "Closing Cash"
    cash_row = b.refs["cf_close"][1]
    data = Reference(b.wb["Cash Flow"], min_col=1, max_col=1 + b.n, min_row=cash_row, max_row=cash_row)
    bar.add_data(data, titles_from_data=True, from_rows=True)
    bar.set_categories(cats)
    bar.height = 8
    bar.width = 18
    ws.add_chart(bar, "B20")


def _checks(b: _Builder):
    ws = b.wb.create_sheet("Checks")
    ws.cell(row=1, column=1, value="Model checks").font = Font(bold=True, color=NAVY, size=12)
    headers = ["Check", "Result", "Status"]
    for j, h in enumerate(headers):
        c = ws.cell(row=3, column=1 + j, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
    last = b.last_col()
    checks = [
        ("Balance sheet balances (max |check|)", f"=MAX(ABS({b.col(0)}{b.refs['bs_check'][1]}:{last}{b.refs['bs_check'][1]}))"),
        ("Cash reconciles (BS cash = CF closing)", f"=SUMPRODUCT(({b.col(0)}{b.refs['bs_cash'][1]}:{last}{b.refs['bs_cash'][1]})-({b.col(0)}{b.refs['cf_close'][1]}:{last}{b.refs['cf_close'][1]}))"),
        ("Minimum cash maintained", f"=MIN({b.col(0)}{b.refs['cf_close'][1]}:{last}{b.refs['cf_close'][1]})-Minimum_Cash_Balance"),
        ("Borrowings non-negative", f"=MIN({b.col(0)}{b.refs['loan_close'][1]}:{last}{b.refs['loan_close'][1]})"),
        ("Projection periods", f"={b.n}"),
    ]
    r = 4
    for label, formula in checks:
        ws.cell(row=r, column=1, value=label).font = F_LABEL
        rc = ws.cell(row=r, column=2, value=formula)
        rc.number_format = CUR
        rc.border = BORDER
        # status: PASS if near zero (for balance/recon) else informational
        sc = ws.cell(row=r, column=3, value=f'=IF(ABS(B{r})<1,"PASS",IF(B{r}<0,"FAIL","WARNING"))')
        sc.border = BORDER
        r += 1
    rng = f"C4:C{r - 1}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"WARNING"'], fill=PatternFill("solid", fgColor=AMBER)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor=RED)))
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12


def _assumptions_summary(b: _Builder):
    ws = b.wb.create_sheet("Assumptions Summary")
    ws.cell(row=1, column=1, value="Assumptions Summary").font = Font(bold=True, color=NAVY, size=12)
    headers = ["Assumption", "Value", "Source sheet"]
    for j, h in enumerate(headers):
        c = ws.cell(row=3, column=1 + j, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
    setup_wc = b.project.working_capital
    tax = b.project.tax
    rows = [
        ("Corporate tax rate", _frac(getattr(tax, "corporate_tax_rate", 0)), "Control Panel", PCT),
        ("VAT rate", _frac(getattr(tax, "vat_rate", 0)), "Control Panel", PCT),
        ("Receivable days", float(getattr(setup_wc, "accounts_receivable_days", 0) or 0), "Working Capital", NUM),
        ("Inventory days", float(getattr(setup_wc, "inventory_days", 0) or 0), "Working Capital", NUM),
        ("Payable days", float(getattr(setup_wc, "accounts_payable_days", 0) or 0), "Working Capital", NUM),
        ("Revenue streams", len(b.project.products), "Revenue Projection", NUM),
        ("Fixed assets", len(b.project.fixed_assets), "Depreciation Schedule", NUM),
        ("Loans", len(b.project.financing.loans) if b.project.financing else 0, "Loan Schedule", NUM),
    ]
    r = 4
    for label, val, src, fmt in rows:
        ws.cell(row=r, column=1, value=label).font = F_LABEL
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = fmt
        c.border = BORDER
        ws.cell(row=r, column=3, value=src).font = Font(color=SLATE)
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24


def _data_dictionary(b: _Builder):
    ws = b.wb.create_sheet("Data Dictionary")
    ws.cell(row=1, column=1, value="Data Dictionary").font = Font(bold=True, color=NAVY, size=12)
    headers = ["Sheet", "Field", "Type", "Description"]
    for j, h in enumerate(headers):
        c = ws.cell(row=3, column=1 + j, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
    rows = [
        ("Control Panel", "Selected Scenario / rates", "Input", "Global model controls (named ranges)."),
        ("Revenue Projection", "Net revenue per stream", "Input", "Driver values; Total = SUM formula."),
        ("Direct/OpEx/Staffing", "Cost lines", "Input", "Driver values; Totals = SUM formulas."),
        ("Depreciation Schedule", "Cost/Residual/Life", "Input", "Straight-line dep + NBV roll (formula)."),
        ("Working Capital", "AR/AP/Inventory days", "Input", "Balances via days formulas."),
        ("Loan Schedule", "Drawdown/rate", "Input", "Equal-principal amortisation (formula)."),
        ("Income Statement", "All lines", "Formula", "Linked to schedules."),
        ("Cash Flow", "All lines", "Formula", "Indirect method; closing cash drives BS cash."),
        ("Balance Sheet", "All lines", "Formula", "Balances by construction; check row = 0."),
        ("Checks", "Reconciliations", "Formula", "PASS/WARNING/FAIL on model integrity."),
    ]
    r = 4
    for s, f, t, d in rows:
        ws.cell(row=r, column=1, value=s).font = F_LABEL
        ws.cell(row=r, column=2, value=f)
        ws.cell(row=r, column=3, value=t)
        ws.cell(row=r, column=4, value=d).font = Font(color=SLATE)
        r += 1
    for col, w in (("A", 22), ("B", 24), ("C", 12), ("D", 52)):
        ws.column_dimensions[col].width = w


def _named_ranges(b: _Builder):
    s, row = b.refs["ctrl_scenario"]
    b.named("Selected_Scenario", s, f"$B${row}")
    b.named("Corporate_Tax_Rate", *_abs(b.refs["ctrl_tax"]))
    b.named("VAT_Rate", *_abs(b.refs["ctrl_vat"]))
    b.named("Minimum_Cash_Balance", *_abs(b.refs["ctrl_mincash"]))


def _abs(ref):
    sheet, row = ref
    return sheet, f"$B${row}"


def _protect(b: _Builder):
    for ws in b.wb.worksheets:
        if ws.title == "Cover":
            continue
        ws.protection.sheet = True
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = False
        ws.protection.autoFilter = False
        ws.protection.formatColumns = True
        ws.protection.formatRows = True
