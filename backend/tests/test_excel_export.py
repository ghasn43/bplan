"""Tests for the Excel financial model export."""
from __future__ import annotations

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.schemas.excel_export import ExcelExportRequest
from app.services import excel_model_service as xl
from app.services.demo_builder import build_demo_project

PID = "demo_aquapure"
B = f"/api/projects/{PID}/exports"

REQUIRED_SHEETS = {
    "Cover", "Control Panel", "Company & Project", "Revenue Projection",
    "Direct Cost Projection", "Operating Expense Projection", "Staffing",
    "Depreciation Schedule", "Working Capital", "Loan Schedule",
    "Income Statement", "Tax & VAT", "Cash Flow", "Balance Sheet",
    "Financial Ratios", "Scenario Comparison", "Charts", "Checks",
    "Assumptions Summary", "Data Dictionary",
}


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        c.post("/api/demo/load-aquapure")
        yield c


@pytest.fixture(scope="module")
def wb():
    proj = build_demo_project()
    f = xl.generate_excel_financial_model(proj, ExcelExportRequest())
    book = openpyxl.load_workbook(xl.export_path(proj.id, f.file_name))
    yield book, f, proj
    import shutil
    shutil.rmtree(xl.EXPORTS_DIR / proj.id, ignore_errors=True)


def _has_formula(ws, max_row=60, max_col=8) -> bool:
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                return True
    return False


# 1. Export route returns 200 with .xlsx + download url
def test_export_route_200(client):
    r = client.post(f"{B}/excel-financial-model", json={"scenario": "base"})
    assert r.status_code == 200
    body = r.json()
    assert body["file_name"].endswith(".xlsx")
    assert body["download_url"].endswith("/download")
    assert body["status"] == "ready"


# 2. Preview returns the sheet list and readiness
def test_preview(client):
    r = client.get(f"{B}/excel-financial-model/preview?scenario=base")
    assert r.status_code == 200
    body = r.json()
    assert body["can_generate"] is True
    assert REQUIRED_SHEETS.issubset(set(body["sheets"]))


# 3. Workbook opens and contains all required sheets
def test_all_sheets_exist(wb):
    book, _f, _p = wb
    assert REQUIRED_SHEETS.issubset(set(book.sheetnames))


# 4-7. Statements + schedules contain formulas
def test_statements_have_formulas(wb):
    book, _f, _p = wb
    assert _has_formula(book["Income Statement"])
    assert _has_formula(book["Balance Sheet"])
    assert _has_formula(book["Cash Flow"])
    assert _has_formula(book["Revenue Projection"])
    assert _has_formula(book["Depreciation Schedule"])
    assert _has_formula(book["Loan Schedule"])


# 8. Income Statement revenue links to the Revenue Projection sheet
def test_is_links_to_schedule(wb):
    book, _f, _p = wb
    isf = book["Income Statement"]
    found = any(
        isinstance(c.value, str) and "Revenue Projection" in c.value
        for row in isf.iter_rows(max_row=40, max_col=8) for c in row
    )
    assert found


# 9. Balance check + cash reconciliation formulas exist
def test_checks_exist(wb):
    book, _f, _p = wb
    bs = book["Balance Sheet"]
    has_check = any(
        isinstance(c.value, str) and c.value.startswith("=") and "-" in c.value
        for row in bs.iter_rows(max_row=40, max_col=8) for c in row
    )
    assert has_check
    checks = book["Checks"]
    assert _has_formula(checks, max_row=20, max_col=4)


# 10. Scenario selector has data validation
def test_scenario_validation(wb):
    book, _f, _p = wb
    cp = book["Control Panel"]
    assert len(cp.data_validations.dataValidation) >= 1


# 11. Named ranges exist
def test_named_ranges(wb):
    book, _f, _p = wb
    names = set(book.defined_names.keys())
    assert {"Selected_Scenario", "Corporate_Tax_Rate", "VAT_Rate", "Minimum_Cash_Balance"} <= names


# 12. Charts are included
def test_charts(wb):
    book, _f, _p = wb
    assert len(book["Charts"]._charts) >= 1


# 13. Company and project names shown separately
def test_company_project_separate(wb):
    book, _f, proj = wb
    cp = book["Company & Project"]
    values = [c.value for row in cp.iter_rows(max_row=20, max_col=2) for c in row]
    assert proj.setup.business_name in values
    assert proj.setup.project_name in values
    assert proj.setup.business_name != proj.setup.project_name


# 14. Input cells unlocked, formula cells locked (protection on by default)
def test_protection(wb):
    book, _f, _p = wb
    rev = book["Revenue Projection"]
    # an input data cell should be unlocked
    assert rev["B4"].protection.locked is False
    assert rev.protection.sheet is True


# 15. Unprotected option leaves the sheet editable
def test_unprotected_option():
    proj = build_demo_project()
    f = xl.generate_excel_financial_model(proj, ExcelExportRequest(protect_formulas=False))
    book = openpyxl.load_workbook(xl.export_path(proj.id, f.file_name))
    assert book["Income Statement"].protection.sheet in (False, None)
    import shutil
    shutil.rmtree(xl.EXPORTS_DIR / proj.id, ignore_errors=True)


# 16. Calculation recalculates on load
def test_calc_on_load(wb):
    book, _f, _p = wb
    assert book.calculation.fullCalcOnLoad is True


# 17. List + download + delete round-trip
def test_list_download_delete(client):
    gen = client.post(f"{B}/excel-financial-model", json={"scenario": "base"}).json()
    assert any(e["export_id"] == gen["export_id"] for e in client.get(B).json())
    dl = client.get(gen["download_url"])
    assert dl.status_code == 200
    assert dl.content[:2] == b"PK"  # xlsx is a zip
    assert client.delete(f"{B}/{gen['export_id']}").status_code == 204


# 18. AquaPure demo workbook generates (company name as entity)
def test_demo_company_name(wb):
    book, _f, proj = wb
    cover_values = [c.value for row in book["Cover"].iter_rows(max_row=10, max_col=4) for c in row]
    assert "AquaPure Smart Filters FZE" in cover_values
